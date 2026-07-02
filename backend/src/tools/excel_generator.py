"""
Excel Generation Tool - create_excel

Creates Excel spreadsheets with structured data.
Based on Anthropic's Agent Skills for xlsx document creation.
"""

import json
import base64
from typing import Annotated, Any
from langchain_core.tools import tool
from langchain_core.runnables import RunnableConfig

from ..runtime_context import user_id_ctx, thread_id_ctx
from ..sandbox import execute_python
from ..storage.file_storage import save_generated_file


def _validate_context(config: RunnableConfig | None) -> tuple[str, str]:
    """Validate and extract user_id and thread_id from config."""
    user_id = None
    thread_id = None
    
    if config:
        configurable = config.get("configurable") or {}
        if isinstance(configurable, dict):
            user_id = configurable.get("user_id")
            thread_id = configurable.get("thread_id")
    
    if not user_id:
        user_id = user_id_ctx.get()
    if not thread_id:
        thread_id = thread_id_ctx.get()
    
    if not user_id or user_id in ("default", "anonymous"):
        raise ValueError("Unable to identify the user. Please sign in again.")
    if not thread_id or thread_id == "default":
        raise ValueError("Unable to identify the thread. Please refresh the page.")
    
    return user_id, thread_id


@tool
def create_excel(
    spec: Annotated[str, """
JSON specification for the Excel workbook:
{
    "sheets": [{
        "name": "Sheet1",
        "headers": ["Column A", "Column B", "Column C"],
        "rows": [["Data 1", 100, 50.5], ["Data 2", 200, 75.3]],
        "formulas": {"D2": "=SUM(B2:C2)"},
        "column_widths": {"A": 20, "B": 15}
    }]
}
    """],
    output_name: Annotated[str, "Output filename, such as 'report.xlsx'"],
    config: RunnableConfig = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """
    Create an Excel workbook.
    
    Suitable for data tables, financial reports, and statistical summaries.
    
    Returns a structured object containing file_id, download_url, and summary.
    """
    try:
        user_id, thread_id = _validate_context(config)
    except ValueError as e:
        return {"success": False, "error": str(e)}
    
    if not output_name.endswith('.xlsx'):
        output_name += '.xlsx'
    
    try:
        spec_dict = json.loads(spec) if isinstance(spec, str) else spec
    except json.JSONDecodeError as e:
        return {"success": False, "error": f"Invalid JSON spec: {e}"}
    
    code = f'''
import json
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

spec = {json.dumps(spec_dict)}

wb = Workbook()

for i, sheet_spec in enumerate(spec.get("sheets", [spec])):
    if i == 0:
        ws = wb.active
        ws.title = sheet_spec.get("name", "Sheet1")
    else:
        ws = wb.create_sheet(title=sheet_spec.get("name", f"Sheet{{i+1}}"))
    
    headers = sheet_spec.get("headers", [])
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    rows = sheet_spec.get("rows", [])
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    formulas = sheet_spec.get("formulas", {{}})
    for cell_ref, formula in formulas.items():
        ws[cell_ref] = formula
    
    widths = sheet_spec.get("column_widths", {{}})
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

import io
buffer = io.BytesIO()
wb.save(buffer)
file_bytes = buffer.getvalue()

print("__FILE_BASE64_START__")
print(base64.b64encode(file_bytes).decode('utf-8'))
print("__FILE_BASE64_END__")
'''
    
    result = execute_python(code=code, timeout=60, user_id=user_id, thread_id=thread_id)
    
    if not result.get("success"):
        return {"success": False, "error": result.get("error") or result.get("stderr", "Execution failed")}
    
    stdout = result.get("stdout", "")
    try:
        start_idx = stdout.index("__FILE_BASE64_START__") + len("__FILE_BASE64_START__")
        end_idx = stdout.index("__FILE_BASE64_END__")
        file_bytes = base64.b64decode(stdout[start_idx:end_idx].strip())
    except Exception as e:
        return {"success": False, "error": f"Failed to retrieve file content: {e}"}
    
    file_info = save_generated_file(
        filename=output_name, data=file_bytes,
        user_id=user_id, thread_id=thread_id, file_type="generated"
    )
    
    if not file_info:
        return {"success": False, "error": "Failed to save file to storage"}
    
    sheets = spec_dict.get("sheets", [spec_dict])
    total_rows = sum(len(s.get("rows", [])) for s in sheets)
    
    return {
        "success": True,
        "file_id": file_info.get("file_id"),
        "download_url": file_info.get("download_url"),
        "filename": output_name,
        "file_type": "generated",
        "summary": f"Created Excel file {output_name} with {len(sheets)} sheets and {total_rows} data rows",
    }


EXCEL_TOOLS = [create_excel]
